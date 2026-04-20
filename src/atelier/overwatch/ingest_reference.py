"""CLI: ingest a reviewer xlsx and emit a curated-reference CSV.

Usage::

    uv run python -m atelier.overwatch.ingest_reference <path.xlsx> \\
        [--out build/data/curated_reference.csv] [--sheet <name>...]

The xlsx is expected to carry one row per reference-labelled column.
The loader scans every row for a short uppercase mnemonic (2–15
chars, alnum + ``_``/``-``) and pairs it with the column name in the
first field.  Multiple sheets may be ingested in one pass.

The emitted CSV has the two-column ``column_name,annotation`` shape.
The mnemonic in the ``annotation`` column (e.g. ``PAN``, ``SSN``,
``EMAIL``) is resolved to a vocabulary code at load time by
:func:`atelier.classify.reference.load_reference_csv`, which consults
``category_set.by_abbrev`` for the lookup.  Mnemonics outside the
loaded vocabulary are logged as unresolved and skipped.

Subsequent pipeline runs that set ``classify_reference_uri`` to this
path will populate ``evaluation_report.json`` with real accuracy
numbers for every column whose mnemonic resolves.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


_DEFAULT_SHEETS = (
    "business_data", "customer_data", "customer_pii_pci_data",
    "Identity_data", "metadata", "transaction_data",
    "personal_data", "system_data",
)


def _is_mnemonic(value: object) -> bool:
    """Short uppercase token — reviewer annotation column shape."""
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not (2 <= len(s) <= 15):
        return False
    if not s.isupper():
        return False
    return s.replace("_", "").replace("-", "").isalnum()


def _extract_rows(wb, sheet_names: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in list(ws.iter_rows(values_only=True))[2:]:
            if not row or not row[0]:
                continue
            col = str(row[0]).strip()
            if not col:
                continue
            tag = next((v.strip() for v in row if _is_mnemonic(v)), None)
            if not tag:
                continue
            key = (col, tag)
            if key in seen:
                continue
            seen.add(key)
            rows.append({"column_name": col, "annotation": tag})
    return rows


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        prog="atelier.overwatch.ingest_reference",
        description="Convert a reviewer xlsx into atelier's curated-reference CSV.",
    )
    ap.add_argument("xlsx", help="Path to the source .xlsx")
    ap.add_argument(
        "--out",
        default="build/data/curated_reference.csv",
        help="Destination CSV path (default build/data/curated_reference.csv)",
    )
    ap.add_argument(
        "--sheet", action="append", default=None,
        help="Sheet name to ingest — repeatable; default covers the 8 reviewer sheets.",
    )
    args = ap.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"error: {xlsx_path} not found", file=sys.stderr)
        return 2

    try:
        import openpyxl  # noqa: WPS433 — top-level import is a runtime dep
    except ImportError:
        print(
            "error: openpyxl is required (uv sync; openpyxl is a runtime dep).",
            file=sys.stderr,
        )
        return 2

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        print(f"error: could not open {xlsx_path}: {exc}", file=sys.stderr)
        return 2

    sheet_names: list[str] = list(args.sheet) if args.sheet else list(_DEFAULT_SHEETS)
    rows = _extract_rows(wb, sheet_names)

    if not rows:
        print(
            f"error: no mnemonics found in sheets {sheet_names!r}",
            file=sys.stderr,
        )
        return 3

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["column_name", "annotation"])
        writer.writeheader()
        writer.writerows(rows)

    print(f"wrote {len(rows)} rows to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
