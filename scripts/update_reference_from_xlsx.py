#!/usr/bin/env python
"""LLM-mediated agent-mediated-reference update from a scoring xlsx.

Reviewers highlight Atelier's predictions in red when they disagree
with the prediction or with the current reference.  Three correction
shapes appear in practice:

1. **Atelier wrong, LLM-solution-column correct** — mechanical extraction
   would suffice, but we route through the LLM for uniformity.
2. **Both Atelier and the LLM-solution-column wrong** — mechanical
   extraction fails; the correct reference is in the reviewer's
   side-note text or implicit in their explanation.
3. **Current reference already correct, flag confirms Atelier was wrong** —
   no reference update needed, but the flag is useful trajectory data.

For every flagged row we hand the LLM concrete trajectory evidence
(sample values, Atelier prediction, LLM-solution tag, reviewer
explanation + side-note, current reference, taxonomy) and ask it to
propose the correct reference annotation with rationale + confidence.

GEPA-shaped: structured trajectory feedback to an LLM that synthesizes
the correct interpretation; output stored as a Pareto-style list of
proposals per row so future iterations can compare candidates.

Modes:
  --mode full    Full 296-entry taxonomy in the prompt (default).
  --mode subset  Structured near-neighbor subset (~15-25 entries).
                 Anchored on Atelier_pred, LLM_sol_tag, current_ref,
                 plus their hierarchical neighborhoods + known
                 confusables from the audit.
  --mode ab      Run both modes on the same rows, write a comparison
                 report.  No --apply when in ab mode.

Output paths (relative to repo root):
  build/data/agent_mediated/
    flagged_rows.json                   # raw extraction from xlsx
    corrections_<xlsx_basename>.json    # all proposals (per-row)
    manual_review_<xlsx_basename>.md    # medium/low-confidence queue
    auto_accept_<xlsx_basename>.json    # high-confidence subset
    ab_comparison_<xlsx_basename>.md    # ab-mode only
    agent_mediated.json.bak             # written on --apply
    changelog.md                        # append-only audit trail

Usage:
  # A/B test on the 5 sample rows
  python scripts/update_reference_from_xlsx.py <xlsx> --mode ab --sample 5

  # Production run, full mode, auto-apply high-confidence
  python scripts/update_reference_from_xlsx.py <xlsx> --mode full --apply

  # Dry run (skip --apply) for inspection
  python scripts/update_reference_from_xlsx.py <xlsx> --mode full
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, "src")

logger = logging.getLogger(__name__)

ANTHROPIC_FALLBACK_MODEL = "claude-opus-4-7"
REFERENCE_PATH = Path("build/data/agent_mediated/agent_mediated.json")
TAXONOMY_CACHE_PATH = Path("build/data/taxonomy/taxonomy_cache.json")

# Hardcoded confusable clusters from the 7bbe4533 audit's top error patterns;
# the subset mode includes these as third-anchor candidates regardless of
# whether they're in the Atelier/LLM neighborhood.
CONFUSABLE_ANCHORS = [
    "0.1",       # INOS
    "1.3.2",     # SYSSTATE
    "1.1.1.1",   # C_FD (financial)
    "1.1.1.1.1", # C_BD (payment)
    "1.2.2",     # TRANSDATE
    "1.2.4",     # TMSTMP
    "1.1.2.2.1.4",  # EMPID
    "1.1.2.2.1.5",  # USERID
    "1.1.2.2.1",    # A_MASID umbrella
    "1.1.1.9.3.1",  # EMAIL
]


# ── xlsx extraction ───────────────────────────────────────────────


# ── Color conventions in scoring xlsx ─────────────────────────────
#
# Reviewers use distinct fill colors to encode different verdicts:
#
#   * Red (FFFF0000) on column D (predicted_annotation) — the prediction
#     is wrong; reviewer expects a correction.  Primary actionable signal.
#   * Light red (FFF4CCCC) on column D — "nominally-incorrect": the
#     prediction is technically wrong but mitigating circumstances apply
#     (taxonomy may not distinguish the cases at the needed granularity,
#     the distinction is genuinely subtle, or there are cross-cutting
#     concerns).  Reviewer wants a taxonomy-review pass, not a quick fix.
#   * Yellow (FFFFFF00) covering the prediction block — row-level review
#     flag with no specific verdict.  Operator should examine.

COLOR_RED = "red"
COLOR_LIGHT_RED = "light_red"
COLOR_YELLOW = "yellow"

_COLOR_BY_RGB = {
    "FFFF0000": COLOR_RED,
    "FFF4CCCC": COLOR_LIGHT_RED,
    "FFFFFF00": COLOR_YELLOW,
}


def _cell_color(cell) -> str | None:
    if not cell.fill or not cell.fill.start_color:
        return None
    rgb = str(cell.fill.start_color.rgb or "")
    return _COLOR_BY_RGB.get(rgb)


def _row_flag(row) -> str | None:
    """Determine the strongest flag color on a row.

    Priority: red on D > light_red on D > yellow on any of A-E.
    Returns None if the row has no flag.
    """
    d_cell = next((c for c in row if c.column_letter == "D"), None)
    if d_cell:
        d_color = _cell_color(d_cell)
        if d_color in (COLOR_RED, COLOR_LIGHT_RED):
            return d_color
    for c in row:
        if c.column_letter in "ABCDE" and _cell_color(c) == COLOR_YELLOW:
            return COLOR_YELLOW
    return None


def extract_flagged_rows(xlsx_path: Path) -> list[dict]:
    """Walk every table sheet, return flagged rows tagged with flag_color
    plus surrounding context.

    Each entry carries ``flag_color`` ∈ {"red", "light_red", "yellow"}.
    The downstream pipeline routes by color: red goes through correction-
    proposal LLM; light_red gets a different system prompt that frames
    the task as taxonomy review; yellow skips LLM and writes to a row-
    review queue for operator examination.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows: list[dict] = []
    for sn in wb.sheetnames:
        if sn == "Overview":
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2):
            flag = _row_flag(row)
            if flag is None:
                continue
            rd = {c.column_letter: c.value for c in row}
            col_name = rd.get("A")
            if not col_name:
                continue
            rows.append({
                "table": sn,
                "column": str(col_name),
                "flag_color": flag,
                "atelier_pred": {
                    "code": rd.get("B"),
                    "label": rd.get("C"),
                    "mnemonic": rd.get("D"),
                    "confidence": rd.get("E"),
                },
                "side_note": rd.get("F"),
                "llm_sol_tag": rd.get("J"),
                "llm_sol_sensitivity": rd.get("K"),
                "llm_sol_explanation": rd.get("M"),
            })
    return rows


# ── reference + taxonomy loading ──────────────────────────────────


def load_reference() -> dict:
    """Load the active agent-mediated reference (qualified → code map)."""
    with open(REFERENCE_PATH) as f:
        return json.load(f)


def load_taxonomy() -> list[dict]:
    """Load the cached taxonomy from default.annotations."""
    if not TAXONOMY_CACHE_PATH.exists():
        sys.exit(
            f"Taxonomy cache missing at {TAXONOMY_CACHE_PATH}; "
            "re-run with --refresh-taxonomy to rebuild."
        )
    with open(TAXONOMY_CACHE_PATH) as f:
        return json.load(f)


def refresh_taxonomy_cache() -> None:
    """Pull fresh taxonomy from hive-poc and persist."""
    import cml.data_v1 as cmldata
    conn = cmldata.get_connection("hive-poc")
    df = conn.get_pandas_dataframe(
        "SELECT id, annotation, definition, common_names, deprecated "
        "FROM default.annotations ORDER BY id"
    )
    out = [
        {
            "code": r["id"],
            "annotation": r["annotation"],
            "definition": r["definition"],
            "common_names": r.get("common_names") or "",
            # Hive's `default.annotations.deprecated` is a string "yes"/"no",
            # not a boolean.  Comparing to "true" here always returned False,
            # silently leaking deprecated codes into the reviewer candidate
            # set at line ~492.  Found via reflect_nhsvm.py 2026-05-25.
            "deprecated": str(r.get("deprecated") or "").strip().lower() == "yes",
        }
        for _, r in df.iterrows()
    ]
    TAXONOMY_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    TAXONOMY_CACHE_PATH.write_text(json.dumps(out, indent=2, default=str))
    print(f"Refreshed {len(out)} taxonomy entries")


def build_code_indexes(taxonomy: list[dict]) -> tuple[dict, dict, dict]:
    """Build (code → entry), (mnemonic → entry), (code → parent_code)."""
    by_code: dict[str, dict] = {t["code"]: t for t in taxonomy}
    by_mnem: dict[str, dict] = {}
    for t in taxonomy:
        mn = (t["annotation"] or "").strip().upper()
        if mn:
            by_mnem.setdefault(mn, t)
    parent_of: dict[str, str | None] = {}
    for t in taxonomy:
        c = t["code"]
        if "." in c:
            parent_of[c] = c.rsplit(".", 1)[0]
        else:
            parent_of[c] = None
    return by_code, by_mnem, parent_of


def resolve_mnemonic_to_code(mnemonic: str | None, by_mnem: dict) -> str | None:
    if not mnemonic:
        return None
    entry = by_mnem.get(mnemonic.strip().upper())
    return entry["code"] if entry else None


# ── subset construction ───────────────────────────────────────────


def near_neighbors(code: str, by_code: dict, parent_of: dict,
                   k_up: int = 2, k_down: int = 1) -> set[str]:
    """Return ``code`` ∪ k_up ancestors ∪ k_down descendants ∪ siblings of ancestors."""
    if code not in by_code:
        return set()
    out = {code}
    # Walk up
    cur = code
    for _ in range(k_up):
        p = parent_of.get(cur)
        if not p:
            break
        out.add(p)
        # Siblings of this ancestor
        for c in by_code:
            if parent_of.get(c) == p:
                out.add(c)
        cur = p
    # Walk down
    def descendants(c: str, depth: int) -> set[str]:
        if depth <= 0:
            return set()
        kids = {x for x in by_code if parent_of.get(x) == c}
        out_ = set(kids)
        for k in kids:
            out_ |= descendants(k, depth - 1)
        return out_
    out |= descendants(code, k_down)
    return out


def build_subset(flagged_row: dict, by_code: dict, by_mnem: dict,
                 parent_of: dict) -> list[dict]:
    """Construct the focused taxonomy subset for one flagged row.

    Anchors: Atelier prediction, LLM-solution tag, current reference,
    plus the hardcoded CONFUSABLE_ANCHORS.  Each anchor contributes
    its k_up=2 / k_down=1 neighborhood.  De-duplicated.
    """
    anchors: list[str] = []
    a_code = flagged_row["atelier_pred"].get("code")
    if a_code and a_code in by_code:
        anchors.append(a_code)
    elif flagged_row["atelier_pred"].get("mnemonic"):
        c = resolve_mnemonic_to_code(
            flagged_row["atelier_pred"]["mnemonic"], by_mnem,
        )
        if c:
            anchors.append(c)
    if flagged_row.get("llm_sol_tag"):
        c = resolve_mnemonic_to_code(flagged_row["llm_sol_tag"], by_mnem)
        if c:
            anchors.append(c)
    if flagged_row.get("current_ref_code"):
        anchors.append(flagged_row["current_ref_code"])
    anchors.extend(CONFUSABLE_ANCHORS)

    codes: set[str] = set()
    for a in anchors:
        codes |= near_neighbors(a, by_code, parent_of)

    return [by_code[c] for c in sorted(codes) if c in by_code]


# ── classification context ────────────────────────────────────────


def load_classifications_lookup(run_dir: Path) -> dict:
    """Map (table, column) → classification row for sample values + ref."""
    cls_path = run_dir / "classifications.json"
    with open(cls_path) as f:
        rows = json.load(f)
    return {(c.get("table_name"), c.get("column_name")): c for c in rows}


# ── prompt construction ──────────────────────────────────────────


# ── Correction-type classification by code relationship ─────────


def correction_type(current_code: str | None, proposed_code: str | None) -> str:
    """Classify the *shape* of a proposed correction by the code
    relationship between current and proposed.

    Returns one of:

    * ``no_current``: current_code missing — can't characterize.
    * ``confirm_current``: same code (the flag was confirmatory).
    * ``granularity_tightening``: proposed is a descendant of current
      (more specific).  Generally safe to auto-apply.
    * ``granularity_loosening``: proposed is an ancestor of current
      (less specific).  Risky — loosening loses information.
    * ``sibling_distinction``: same parent, different leaf.  Often
      taxonomy-subtle; review preferred.
    * ``subtree_correction``: different branch entirely.  The most
      decisive correction shape.
    """
    if not current_code or not proposed_code:
        return "no_current"
    if current_code == proposed_code:
        return "confirm_current"
    if proposed_code.startswith(current_code + "."):
        return "granularity_tightening"
    if current_code.startswith(proposed_code + "."):
        return "granularity_loosening"
    cur_parent = current_code.rsplit(".", 1)[0] if "." in current_code else ""
    pro_parent = proposed_code.rsplit(".", 1)[0] if "." in proposed_code else ""
    if cur_parent and cur_parent == pro_parent:
        return "sibling_distinction"
    return "subtree_correction"


# ── System prompts ────────────────────────────────────────────────


SYSTEM_PROMPT = """\
You are an ontology curation reviewer for a data classification system.
A human reviewer has flagged a column's reference annotation as wrong.
Your job is to propose the correct annotation given concrete evidence.

You will receive:
* Column identity (table, column name) and sample values.
* What the Atelier classifier predicted (flagged as wrong).
* What an upstream "LLM solution" tagged it (sometimes blank).
* The reviewer's free-text explanation and a side-note.
* The current agent-mediated reference assignment for this column.
* A taxonomy list (code, mnemonic, definition).

Pick the correct mnemonic by reasoning over the evidence:
* Sample values dominate over column name when they conflict.
* The reviewer's side-note often contains the intended mnemonic
  explicitly or names a concept that maps to one.
* If the current reference is already correct (Atelier was just wrong),
  proposed_mnemonic should equal the current reference's mnemonic and
  current_assessment should be "current_reference_is_correct".
* If neither Atelier nor the upstream LLM is right and the side-note
  is uninformative, you may pick a third option — but only if the
  evidence clearly supports it.  Otherwise say insufficient_evidence.

Output STRICT JSON only:
{
  "proposed_mnemonic": "<mnemonic from taxonomy>" | "INSUFFICIENT_EVIDENCE",
  "confidence": "high" | "medium" | "low",
  "reasoning": "<2-4 sentences>",
  "alternatives_considered": ["<mnemonic> — <why rejected>"],
  "current_assessment": "current_reference_is_wrong" | "current_reference_is_correct" | "ambiguous"
}

No markdown, no preamble.
"""


SYSTEM_PROMPT_LIGHT_RED = """\
You are an ontology curation reviewer evaluating a "nominally-incorrect"
classification.  The human reviewer flagged the prediction in LIGHT RED
rather than dark red — meaning: the current prediction is technically
wrong, BUT the reviewer signaled mitigating circumstances.  Typically
one of:

* The taxonomy may not be granular enough to distinguish the correct
  case (e.g. the reviewer's source taxonomy has a finer distinction
  that doesn't exist here).
* The semantic distinction between the predicted code and the "right"
  code is genuinely subtle (e.g. sibling-leaf disambiguation under
  the same parent, where both could be defended).
* There are cross-cutting concerns that make a confident pick
  impractical without more domain context.

Your task:
1. Diagnose whether the current taxonomy CAN distinguish the right
   classification at the appropriate granularity.
2. If yes, propose the correct mnemonic with reasoning.
3. If no, set ``proposed_mnemonic`` to ``TAXONOMY_GAP`` and explain
   what's missing in the ``taxonomy_observation`` field — concrete
   enough that a curator can act on it.

Output STRICT JSON only:
{
  "proposed_mnemonic": "<mnemonic>" | "TAXONOMY_GAP" | "INSUFFICIENT_EVIDENCE",
  "confidence": "high" | "medium" | "low",
  "reasoning": "<2-4 sentences>",
  "taxonomy_observation": "<what's missing or what makes this subtle, "
                         "or empty if taxonomy is sufficient>",
  "alternatives_considered": ["<mnemonic> — <why rejected>"],
  "current_assessment": "current_reference_is_wrong" | "current_reference_is_correct" | "ambiguous"
}

No markdown, no preamble.
"""


def build_user_prompt(
    flagged: dict,
    taxonomy_subset: list[dict],
    by_code: dict,
) -> str:
    """Compose the per-row prompt body."""
    lines = []
    lines.append(f"# Column under review")
    lines.append(f"- Table: `{flagged['table']}`")
    lines.append(f"- Column: `{flagged['column']}`")
    if flagged.get("embedding_text"):
        lines.append(f"- Sample/embedding text: \"{flagged['embedding_text'][:400]}\"")

    a = flagged["atelier_pred"]
    lines.append(f"\n# Atelier prediction (flagged wrong by reviewer)")
    lines.append(f"- mnemonic: `{a.get('mnemonic')}`")
    lines.append(f"- code: `{a.get('code')}`")
    lines.append(f"- label: \"{a.get('label')}\"")
    lines.append(f"- confidence: {a.get('confidence')}")

    lines.append(f"\n# Upstream LLM-solution tag")
    lines.append(f"- tag: `{flagged.get('llm_sol_tag') or '<blank>'}`")
    lines.append(f"- sensitivity: `{flagged.get('llm_sol_sensitivity') or '<blank>'}`")
    lines.append(f"- explanation: \"{flagged.get('llm_sol_explanation') or ''}\"")

    if flagged.get("side_note"):
        lines.append(f"\n# Reviewer side-note")
        lines.append(f"\"{flagged['side_note']}\"")

    cur_code = flagged.get("current_ref_code")
    cur_entry = by_code.get(cur_code) if cur_code else None
    lines.append(f"\n# Current agent-mediated reference")
    if cur_entry:
        lines.append(
            f"- code: `{cur_code}`  mnemonic: `{cur_entry['annotation']}`"
        )
        lines.append(f"- definition: \"{cur_entry['definition']}\"")
    else:
        lines.append(f"- code: `{cur_code or '<unset>'}` (not in taxonomy)")

    lines.append(f"\n# Taxonomy ({len(taxonomy_subset)} entries)")
    # Deprecated codes remain in the list — deprecation is a curation
    # signal (prefer non-deprecated alternatives), NOT a hide.  Columns
    # already tagged with a deprecated code may legitimately keep that
    # assignment; hiding the code from the reviewer would force a
    # spurious re-routing.
    for t in taxonomy_subset:
        suffix = "  [DEPRECATED]" if t["deprecated"] else ""
        line = f"- `{t['code']}` `{t['annotation']}`{suffix} — {t['definition'][:160]}"
        if t.get("common_names"):
            line += f"  (aliases: {t['common_names'][:100]})"
        lines.append(line)

    return "\n".join(lines)


# ── LLM call ─────────────────────────────────────────────────────


def resolve_model(cfg, override: str | None = None) -> str:
    if override:
        return override
    candidate = (getattr(cfg, "agent_model", "") or "").strip()
    return candidate or ANTHROPIC_FALLBACK_MODEL


def call_llm(system: str, user: str, cfg, model: str) -> dict:
    """Make one Anthropic call, parse strict-JSON response."""
    from atelier.agents.client import (
        _build_anthropic_client, _build_bedrock_client,
    )
    is_bedrock = model.startswith("arn:") or "anthropic." in model
    client = (
        _build_bedrock_client(cfg, timeout=120.0) if is_bedrock
        else _build_anthropic_client(cfg, timeout=120.0)
    )
    resp = client.messages.create(
        model=model,
        max_tokens=1024,
        system=[{"type": "text", "text": system,
                 "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": user}],
    )
    text = "".join(
        b.text for b in resp.content if hasattr(b, "text")
    ).strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.strip().rstrip("`")
    return json.loads(text)


# ── apply step ────────────────────────────────────────────────────


def _entry_mnemonic(v: object) -> str:
    """Extract the mnemonic from a reference entry (dual or legacy)."""
    if isinstance(v, dict):
        return (v.get("mnemonic") or "").strip()
    if isinstance(v, str):
        return v.strip()
    return ""


def _entry_code(v: object) -> str | None:
    """Extract the captured code from a reference entry (dual only)."""
    if isinstance(v, dict):
        return v.get("code")
    return None


def apply_corrections(
    reference: dict,
    accepted: list[dict],
    by_mnem: dict,
    xlsx_basename: str,
) -> tuple[dict, list[dict]]:
    """Apply high-confidence corrections to the reference dict.

    Writes entries in **dual format**: each value is a dict with
    ``mnemonic``, ``code``, ``captured_at``, and ``source`` fields.
    This pairs the human-readable mnemonic with the hierarchical code
    at capture time so that future mnemonic→code drift (curator moves
    a mnemonic to a different code/subtree) becomes detectable at load
    time rather than silently changing the entry's meaning.

    Handles legacy mnemonic-string entries on read (for the no-op
    check); always emits dual entries on write.

    Returns (updated_reference, changelog_entries).  Caller is
    responsible for snapshot + file write.
    """
    updated = dict(reference)
    changes = []
    now_iso = datetime.now(timezone.utc).isoformat()
    source_tag = f"xlsx:{xlsx_basename}"

    for corr in accepted:
        if corr.get("status") != "high_apply":
            continue
        mn = corr["proposed_mnemonic"]
        if mn in ("INSUFFICIENT_EVIDENCE", None):
            continue
        new_entry = by_mnem.get(mn.strip().upper())
        if not new_entry:
            corr.setdefault("apply_skip_reason", f"unknown mnemonic {mn!r}")
            continue
        new_mnemonic = (new_entry["annotation"] or "").strip()
        new_code = new_entry["code"]
        qkey = f"{corr['table']}.{corr['column']}"
        old_value = updated.get(qkey)
        old_mnemonic = _entry_mnemonic(old_value)
        old_code = _entry_code(old_value)
        if old_mnemonic.upper() == new_mnemonic.upper():
            corr["apply_skip_reason"] = "no-op (already correct mnemonic)"
            continue
        updated[qkey] = {
            "mnemonic": new_mnemonic,
            "code": new_code,
            "captured_at": now_iso,
            "source": source_tag,
        }
        changes.append({
            "qkey": qkey,
            "old_mnemonic": old_mnemonic or None,
            "old_code": old_code,
            "new_mnemonic": new_mnemonic,
            "new_code": new_code,
            "confidence": corr.get("confidence"),
            "reasoning": corr.get("reasoning"),
            "source": source_tag,
            "applied_at": now_iso,
        })
    return updated, changes


def write_changelog(changes: list[dict]) -> None:
    """Append-only changelog at build/data/agent_mediated/changelog.md."""
    path = Path("build/data/agent_mediated/changelog.md")
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    if not path.exists():
        lines.append("# Agent-mediated reference changelog")
        lines.append("")
    lines.append(
        f"\n## {datetime.now(timezone.utc).isoformat()} — "
        f"{len(changes)} corrections applied"
    )
    lines.append("")
    if not changes:
        lines.append("(no-op)")
    else:
        lines.append("| qkey | old | new | code | conf |")
        lines.append("|---|---|---|---|---|")
        for c in changes:
            lines.append(
                f"| `{c['qkey']}` | `{c['old_mnemonic']}` | "
                f"`{c['new_mnemonic']}` | `{c['new_code']}` | "
                f"{c['confidence']} |"
            )
    with path.open("a") as f:
        f.write("\n".join(lines) + "\n")


# ── per-row pipeline ──────────────────────────────────────────────


def enrich_row(
    flagged: dict,
    cls_lookup: dict,
    reference: dict,
) -> dict:
    """Attach current_ref_code + embedding_text from classifications."""
    cls = cls_lookup.get((flagged["table"], flagged["column"]))
    if cls:
        flagged["embedding_text"] = cls.get("embedding_text", "")
        flagged["current_ref_code"] = cls.get("reference_code")
    if not flagged.get("current_ref_code"):
        qkey = f"{flagged['table']}.{flagged['column']}"
        flagged["current_ref_code"] = reference.get(qkey)
    return flagged


def process_row(
    flagged: dict,
    taxonomy: list[dict],
    by_code: dict, by_mnem: dict, parent_of: dict,
    mode: str,
    cfg, model: str,
) -> dict:
    """Branch on flag_color and call LLM in the requested taxonomy-
    presentation mode(s).  Returns enriched proposal record.

    * **red** → standard correction prompt (SYSTEM_PROMPT).
    * **light_red** → taxonomy-review prompt (SYSTEM_PROMPT_LIGHT_RED)
      that allows the LLM to flag TAXONOMY_GAP rather than forcing
      a mnemonic pick.
    * **yellow** → no LLM call; the row is captured as a review-queue
      observation for the operator.
    """
    base = {
        "table": flagged["table"],
        "column": flagged["column"],
        "flag_color": flagged.get("flag_color"),
        "current_ref_code": flagged.get("current_ref_code"),
        "atelier_pred": flagged["atelier_pred"],
        "llm_sol_tag": flagged.get("llm_sol_tag"),
        "side_note": flagged.get("side_note"),
    }

    flag = flagged.get("flag_color")
    if flag == COLOR_YELLOW:
        # No LLM call — yellow is "operator notice", not a correction.
        return base

    sys_prompt = (
        SYSTEM_PROMPT_LIGHT_RED if flag == COLOR_LIGHT_RED else SYSTEM_PROMPT
    )

    if mode in ("full", "ab"):
        prompt = build_user_prompt(flagged, taxonomy, by_code)
        try:
            base["full"] = call_llm(sys_prompt, prompt, cfg, model)
        except Exception as exc:  # noqa: BLE001
            base["full"] = {"error": f"{type(exc).__name__}: {exc}"}

    if mode in ("subset", "ab"):
        subset = build_subset(flagged, by_code, by_mnem, parent_of)
        prompt = build_user_prompt(flagged, subset, by_code)
        try:
            base["subset"] = call_llm(sys_prompt, prompt, cfg, model)
        except Exception as exc:  # noqa: BLE001
            base["subset"] = {"error": f"{type(exc).__name__}: {exc}"}
        base["subset_size"] = len(subset)

    return base


def classify_proposal(
    record: dict, mode: str, by_mnem: dict,
) -> tuple[str, dict]:
    """Decide apply-status for a proposal record.

    Status taxonomy:

    * ``row_review`` — yellow flag, no LLM call, operator notice only.
    * ``taxonomy_review`` — light_red flag.  Never auto-apply; the
      proposal becomes a taxonomy-curation candidate.  TAXONOMY_GAP
      results route here explicitly.
    * ``high_apply`` — red + high confidence + subtree_correction OR
      granularity_tightening.
    * ``manual_review`` — red + (low/medium confidence) OR red + high
      confidence on a sibling_distinction / granularity_loosening
      (these are reference-altering moves that deserve human eyes).
    * ``confirm_current`` — LLM verdict matches current reference.
    * ``insufficient_evidence`` — LLM declined to propose.
    * ``error`` — LLM call failed.

    Returns (status, primary_proposal_dict).
    """
    flag = record.get("flag_color")

    if flag == COLOR_YELLOW:
        return ("row_review", {})

    primary = record.get(mode if mode != "ab" else "full") or {}
    if "error" in primary:
        return ("error", primary)

    mn = primary.get("proposed_mnemonic")
    conf = primary.get("confidence")

    if mn in ("INSUFFICIENT_EVIDENCE", None):
        return ("insufficient_evidence", primary)
    if mn == "TAXONOMY_GAP":
        return ("taxonomy_review", primary)
    if primary.get("current_assessment") == "current_reference_is_correct":
        return ("confirm_current", primary)

    # Compute correction shape for routing
    cur_code = record.get("current_ref_code")
    pro_entry = by_mnem.get((mn or "").strip().upper())
    pro_code = pro_entry["code"] if pro_entry else None
    ctype = correction_type(cur_code, pro_code)
    primary["correction_type"] = ctype

    if flag == COLOR_LIGHT_RED:
        # Light red never auto-applies; goes to taxonomy review even
        # when a concrete mnemonic was proposed.  The proposal is
        # surfaced as a curator-actionable suggestion.
        return ("taxonomy_review", primary)

    if conf == "high":
        if ctype in ("subtree_correction", "granularity_tightening"):
            return ("high_apply", primary)
        # sibling_distinction / granularity_loosening / no_current
        # deserve human eyes even at high confidence.
        return ("manual_review", primary)

    return ("manual_review", primary)


# ── CLI ──────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("xlsx", help="Path to scoring xlsx with red highlights")
    parser.add_argument(
        "--mode", choices=["full", "subset", "ab"], default="full",
        help="Taxonomy presentation mode (default: full)",
    )
    parser.add_argument(
        "--sample", type=int, default=0,
        help="Limit to first N flagged rows (0 = all)",
    )
    parser.add_argument(
        "--sample-cols", default=None,
        help="Comma-separated 'table.column' list to restrict to "
             "(supersedes --sample)",
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Apply high-confidence corrections to agent_mediated.json. "
             "Snapshots .bak first; appends to changelog.md.",
    )
    parser.add_argument(
        "--run-dir", default="build/results/7bbe4533",
        help="Run dir for classifications.json context",
    )
    parser.add_argument(
        "--model", default=None,
        help="LLM model override (default resolves from cfg.agent_model)",
    )
    parser.add_argument(
        "--refresh-taxonomy", action="store_true",
        help="Refresh the cached taxonomy from default.annotations first",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")
    xlsx_basename = xlsx_path.stem

    if args.refresh_taxonomy or not TAXONOMY_CACHE_PATH.exists():
        refresh_taxonomy_cache()

    print(f"Loading taxonomy from {TAXONOMY_CACHE_PATH}...")
    taxonomy = load_taxonomy()
    by_code, by_mnem, parent_of = build_code_indexes(taxonomy)
    print(f"  {len(taxonomy)} entries indexed")

    print(f"Extracting flagged rows from {xlsx_path}...")
    flagged_rows = extract_flagged_rows(xlsx_path)
    print(f"  {len(flagged_rows)} flagged rows")

    print(f"Loading reference + classifications context...")
    reference = load_reference()
    cls_lookup = load_classifications_lookup(Path(args.run_dir))
    for f in flagged_rows:
        enrich_row(f, cls_lookup, reference)

    out_dir = REFERENCE_PATH.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "flagged_rows.json").write_text(
        json.dumps(flagged_rows, indent=2, default=str),
    )

    # Apply sample filter
    if args.sample_cols:
        wanted = set(args.sample_cols.split(","))
        flagged_rows = [
            f for f in flagged_rows
            if f"{f['table']}.{f['column']}" in wanted
        ]
        print(f"  filtered to {len(flagged_rows)} via --sample-cols")
    elif args.sample > 0:
        flagged_rows = flagged_rows[:args.sample]
        print(f"  limited to first {len(flagged_rows)} via --sample")

    print(f"Loading config + model...")
    from atelier.config import load_config
    cfg = load_config()
    model = resolve_model(cfg, args.model)
    print(f"  model: {model}")

    if args.mode == "ab" and args.apply:
        print("WARNING: --apply ignored in --mode ab; ab mode is research-only")
        args.apply = False

    print(f"\nProcessing {len(flagged_rows)} row(s) in mode={args.mode}...")
    records = []
    for i, f in enumerate(flagged_rows, 1):
        qkey = f"{f['table']}.{f['column']}"
        print(f"  [{i}/{len(flagged_rows)}] {qkey}")
        rec = process_row(
            f, taxonomy, by_code, by_mnem, parent_of,
            args.mode, cfg, model,
        )
        status, primary = classify_proposal(rec, args.mode, by_mnem)
        rec["status"] = status
        rec["proposed_mnemonic"] = primary.get("proposed_mnemonic")
        rec["confidence"] = primary.get("confidence")
        rec["reasoning"] = primary.get("reasoning")
        records.append(rec)
        print(
            f"    → {status}  proposed={rec['proposed_mnemonic']!r}  "
            f"conf={rec['confidence']!r}"
        )

    # Persist
    corrections_path = out_dir / f"corrections_{xlsx_basename}.json"
    corrections_path.write_text(json.dumps({
        "xlsx": str(xlsx_path),
        "mode": args.mode,
        "model": model,
        "run_at": datetime.now(timezone.utc).isoformat(),
        "n_flagged": len(flagged_rows),
        "records": records,
    }, indent=2, default=str))
    print(f"\nWrote {corrections_path}")

    # Status summary
    status_counts = Counter(r["status"] for r in records)
    print(f"\nStatus distribution:")
    for st, n in status_counts.most_common():
        print(f"  {st}: {n}")

    # Manual review queue (red-flag rows that didn't reach high-confidence
    # auto-apply: low/medium confidence, sibling/loosening at high conf,
    # insufficient_evidence, or LLM error).
    review = [
        r for r in records
        if r["status"] in ("manual_review", "insufficient_evidence", "error")
    ]
    if review:
        review_path = out_dir / f"manual_review_{xlsx_basename}.md"
        lines = [f"# Manual review queue — {xlsx_basename}", ""]
        for r in review:
            primary = r.get(args.mode if args.mode != "ab" else "full") or {}
            ctype = primary.get("correction_type", "—")
            lines.append(f"## {r['table']}.{r['column']}  ({r['status']})")
            lines.append(f"- Atelier: `{r['atelier_pred'].get('mnemonic')}` "
                         f"({r['atelier_pred'].get('code')})")
            lines.append(f"- LLM-sol: `{r.get('llm_sol_tag')}`")
            lines.append(f"- Side-note: {r.get('side_note')}")
            lines.append(f"- Proposed: `{r['proposed_mnemonic']}`  "
                         f"conf=`{r['confidence']}`  type=`{ctype}`")
            lines.append(f"- Reasoning: {r['reasoning']}")
            lines.append("")
        review_path.write_text("\n".join(lines))
        print(f"Wrote {review_path}")

    # Taxonomy review queue (light_red rows + TAXONOMY_GAP results).
    # These are curator-actionable suggestions, not pipeline corrections.
    tax = [r for r in records if r["status"] == "taxonomy_review"]
    if tax:
        tax_path = out_dir / f"taxonomy_review_{xlsx_basename}.md"
        lines = [
            f"# Taxonomy review queue — {xlsx_basename}",
            "",
            "Light-red flags ('nominally-incorrect') or LLM-declared "
            "TAXONOMY_GAP results.  These are taxonomy-curator-actionable "
            "items, not reference corrections.  The current prediction is "
            "wrong but mitigating circumstances apply — review the "
            "taxonomy_observation for what's missing.",
            "",
        ]
        for r in tax:
            primary = r.get(args.mode if args.mode != "ab" else "full") or {}
            obs = primary.get("taxonomy_observation", "")
            ctype = primary.get("correction_type", "—")
            lines.append(f"## {r['table']}.{r['column']}")
            lines.append(f"- Atelier: `{r['atelier_pred'].get('mnemonic')}` "
                         f"({r['atelier_pred'].get('code')})")
            lines.append(f"- LLM-sol: `{r.get('llm_sol_tag')}`")
            lines.append(f"- Side-note: {r.get('side_note')}")
            lines.append(f"- Proposed: `{r['proposed_mnemonic']}`  "
                         f"conf=`{r['confidence']}`  type=`{ctype}`")
            lines.append(f"- Reasoning: {r['reasoning']}")
            if obs:
                lines.append(f"- Taxonomy observation: {obs}")
            lines.append("")
        tax_path.write_text("\n".join(lines))
        print(f"Wrote {tax_path}")

    # Row-review queue (yellow flags — operator-notice rows; no LLM call,
    # no corrections proposed).
    yellow = [r for r in records if r["status"] == "row_review"]
    if yellow:
        yellow_path = out_dir / f"row_review_{xlsx_basename}.md"
        lines = [
            f"# Row review queue — {xlsx_basename}",
            "",
            "Reviewer flagged these rows for examination without a "
            "specific verdict.  Useful as adjacent context for nearby "
            "corrections and for spotting categorically problematic "
            "rows; no pipeline action is taken.",
            "",
        ]
        for r in yellow:
            lines.append(f"## {r['table']}.{r['column']}")
            lines.append(f"- Atelier: `{r['atelier_pred'].get('mnemonic')}` "
                         f"({r['atelier_pred'].get('code')})")
            lines.append(f"- LLM-sol: `{r.get('llm_sol_tag')}`")
            lines.append(f"- Side-note: {r.get('side_note')}")
            lines.append(f"- LLM-sol explanation: {r.get('llm_sol_explanation')}")
            lines.append(f"- Current ref: `{r.get('current_ref_code')}`")
            lines.append("")
        yellow_path.write_text("\n".join(lines))
        print(f"Wrote {yellow_path}")

    # AB comparison report
    if args.mode == "ab":
        report_path = out_dir / f"ab_comparison_{xlsx_basename}.md"
        lines = [f"# A/B comparison — {xlsx_basename}", "", f"Model: `{model}`", ""]
        agree = 0
        for r in records:
            full = (r.get("full") or {}).get("proposed_mnemonic")
            sub = (r.get("subset") or {}).get("proposed_mnemonic")
            if full == sub:
                agree += 1
            lines.append(f"## {r['table']}.{r['column']}")
            lines.append(f"- side-note: {r.get('side_note')}")
            lines.append(f"- current_ref: `{r.get('current_ref_code')}`")
            lines.append(f"- full mode:   `{full}`  (subset={r.get('subset_size','?')})")
            lines.append(f"  reasoning: {(r.get('full') or {}).get('reasoning')}")
            lines.append(f"- subset mode: `{sub}`")
            lines.append(f"  reasoning: {(r.get('subset') or {}).get('reasoning')}")
            same = "✅" if full == sub else "⚠️"
            lines.append(f"- agreement: {same}")
            lines.append("")
        lines.insert(2, f"\nFull/subset agreement: {agree}/{len(records)}")
        report_path.write_text("\n".join(lines))
        print(f"Wrote {report_path}")

    # Auto-apply (only when not in ab mode)
    if args.apply:
        accepted = [r for r in records if r["status"] == "high_apply"]
        print(f"\n--apply: {len(accepted)} high-confidence correction(s) to apply")
        updated, changes = apply_corrections(
            reference, accepted, by_mnem, xlsx_basename,
        )
        if changes:
            backup_path = REFERENCE_PATH.with_suffix(".json.bak")
            shutil.copy2(REFERENCE_PATH, backup_path)
            print(f"  Snapshot → {backup_path}")
            REFERENCE_PATH.write_text(json.dumps(updated, indent=2, default=str))
            print(f"  Wrote updated reference → {REFERENCE_PATH}")
            write_changelog(changes)
            print(f"  Appended changelog entries: {len(changes)}")
        else:
            print(f"  No effective changes (proposals matched current reference)")

    print(f"\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
