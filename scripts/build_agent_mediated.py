#!/usr/bin/env python
"""Assemble the per-column working set for agent-mediated reference review.

This script does NO classification.  It only gathers inputs so that an
independent reviewer (this session's Opus 4.7 + ultrathink, per the
project's agent-mediated reference principles memory) can produce a defensible
agent_mediated.json — see ``project_agent_mediated_principles.md``.

Inputs gathered per column
--------------------------

* xlsx row: column_name, Atelier prediction (code, label, annotation,
  confidence), and the competing LLM-prompt-solution's annotation +
  sensitivity + confidence + explanation.
* Latest classifications.json row: ``predicted_annotation``, belief,
  conflict, evidence_sources breakdown, and ``embedding_text`` (which
  carries the column's sample values).
* Vocabulary: ``default.annotations`` rows indexed by mnemonic so the
  reviewer can quickly look up valid tags + definitions.

Output
------

``build/data/agent_mediated/working_set.json`` shaped as::

  {
    "metadata": {
      "built_at": "...",
      "xlsx_path": "...",
      "run_id": "981f69be",
      "vocabulary_size": 296,
      "table_count": 20,
      "column_count": 304
    },
    "vocabulary": {
      "INOS": {"code": "0.1", "label": "Internal Non-Sensitive", "definition": "..."},
      ...
    },
    "columns": {
      "legal_cases.row_id": {
        "table_name": "legal_cases",
        "column_name": "row_id",
        "xlsx": {
          "atelier": {"code": "0.1", "label": "...", "annotation": "INOS", "confidence": 0.822},
          "llm_solution": {"annotation": null, "sensitivity": "public", "confidence": "high",
                           "explanation": "Appears to be a technical row identifier..."}
        },
        "atelier_current": {
          "run_id": "981f69be",
          "predicted_annotation": "INOS",
          "predicted_code": "0.1",
          "belief": 0.8412,
          "conflict": 0.3267,
          "evidence_text": "row_id | int64 | 1, 2, 3, 4, 5 | cardinality=1000 | ..."
        }
      },
      ...
    }
  }

Then the reviewer (this Opus session) walks ``columns`` table-by-table
and writes the agent-mediated reference + audit artifacts.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path


def _utc_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _load_vocabulary(connection_name: str, database: str) -> dict:
    """Pull the vocabulary table and index by mnemonic.

    Returns ``{annotation_mnemonic: {code, label, definition, deprecated}}``.
    """
    import cml.data_v1 as cmldata
    conn = cmldata.get_connection(connection_name)
    try:
        df = conn.get_pandas_dataframe(
            f"SELECT id, ontology, annotation, definition, deprecated "
            f"FROM {database}.annotations"
        )
    finally:
        conn.close()

    # Hive returns columns prefixed by table name; strip it.
    df.columns = [c.split(".")[-1] for c in df.columns]

    vocab = {}
    for _, row in df.iterrows():
        mnem = str(row.get("annotation") or "").strip()
        if not mnem:
            continue
        vocab[mnem] = {
            "code": str(row.get("id") or "").strip(),
            "label": str(row.get("ontology") or "").strip(),
            "definition": str(row.get("definition") or "").strip(),
            "deprecated": str(row.get("deprecated") or "").strip().lower() == "yes",
        }
    return vocab


def _parse_xlsx(xlsx_path: Path) -> dict:
    """Parse the comparison xlsx into ``{table.col: xlsx_entry}``.

    Tolerates header drift (one sheet uses 'e' instead of 'column_name'
    in header row 0 — we still take col 0 as the column name).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "overview":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Header row 0 indicates where the LLM-solution section begins.
        # Find 'Asset' column index; LLM-side columns follow at fixed
        # offsets: Asset, Column Name, Suggested Classification Tags,
        # Data Sensitivity, Confidence Score, Explanation.  The number
        # of spacer Nones between the Atelier section and 'Asset' varies
        # by sheet (1 or 2), so we anchor to 'Asset' rather than to a
        # hardcoded offset.
        header = rows[0]
        asset_idx = None
        for i, v in enumerate(header):
            if v and str(v).strip().lower() == "asset":
                asset_idx = i
                break
        llm_tag_idx = (asset_idx + 2) if asset_idx is not None else None
        llm_sens_idx = (asset_idx + 3) if asset_idx is not None else None
        llm_conf_idx = (asset_idx + 4) if asset_idx is not None else None
        llm_expl_idx = (asset_idx + 5) if asset_idx is not None else None

        def _cell(row, idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            v = str(row[idx]).strip()
            return v or None

        for row in rows[1:]:
            if not row or row[0] is None:
                break  # blank row terminates the column section
            column_name = str(row[0]).strip()
            if not column_name:
                break
            # Some sheets repeat the header partway down as a separator
            # before reviewer notes; stop parsing column data there.
            if column_name.lower() in {"column_name", "column name"}:
                break
            # The predicted_code cell should be numeric-ish (e.g. "0.1");
            # if it's the literal "predicted_code" header, we've fallen off
            # the column section into a note section.
            if isinstance(row[1], str) and row[1].lower() in {
                "predicted_code", "predicted code",
            }:
                break
            # Atelier's prediction columns (cols 0-4)
            atelier_code = row[1]
            atelier_code = str(atelier_code).strip() if atelier_code is not None else None
            atelier_entry = {
                "code": atelier_code,
                "label": str(row[2]).strip() if row[2] else None,
                "annotation": str(row[3]).strip() if row[3] else None,
                "confidence": float(row[4]) if row[4] is not None else None,
            }
            # LLM-prompt-solution columns: indices computed from 'Asset' header.
            llm_entry = {
                "annotation":  _cell(row, llm_tag_idx),
                "sensitivity": _cell(row, llm_sens_idx),
                "confidence":  _cell(row, llm_conf_idx),
                "explanation": _cell(row, llm_expl_idx),
            }
            out[f"{sheet_name}.{column_name}"] = {
                "table_name": sheet_name,
                "column_name": column_name,
                "atelier": atelier_entry,
                "llm_solution": llm_entry,
            }
    return out


def _load_classifications(run_dir: Path) -> dict:
    """Return ``{table.col: classifications.json row}``."""
    path = run_dir / "classifications.json"
    if not path.exists():
        return {}
    rows = json.loads(path.read_text())
    out = {}
    for row in rows:
        t = str(row.get("table_name") or "").strip()
        c = str(row.get("column_name") or "").strip()
        if not t or not c:
            continue
        out[f"{t}.{c}"] = row
    return out


def _latest_run_dir(results_root: Path) -> Path | None:
    candidates = [p for p in results_root.iterdir() if p.is_dir()]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


# Some xlsx sheet names were typos / contained spaces that have since
# been corrected in the live Hive table names.  Map xlsx → Hive so we
# key the working set by the Hive name that the sweep runs will produce.
XLSX_NAME_FIXES = {
    "hr_compenstation": "hr_compensation",
    "shipping manifests": "shipping_manifests",
}


def _list_hive_tables(connection_name: str, database: str) -> list[str]:
    import cml.data_v1 as cmldata
    conn = cmldata.get_connection(connection_name)
    try:
        df = conn.get_pandas_dataframe(f"SHOW TABLES IN {database}")
    finally:
        conn.close()
    # Column may be 'tab_name' or '{database}.tab_name' depending on dialect.
    col = df.columns[0]
    return sorted(str(v).strip() for v in df[col].tolist())


def _sample_hive_columns(
    connection_name: str,
    database: str,
    table: str,
    n_samples: int = 10,
) -> list[dict]:
    """Pull column names, dtypes, and up to ``n_samples`` sample values."""
    import cml.data_v1 as cmldata
    conn = cmldata.get_connection(connection_name)
    try:
        df = conn.get_pandas_dataframe(
            f"SELECT * FROM {database}.{table} LIMIT {n_samples}"
        )
    finally:
        conn.close()
    cols = []
    for raw_col in df.columns:
        # Hive returns columns prefixed by table name; strip it.
        col = str(raw_col).split(".")[-1]
        series = df[raw_col]
        samples = []
        for v in series.tolist():
            if v is None:
                continue
            s = str(v).strip()
            if not s or s == "None":
                continue
            samples.append(s[:120])  # truncate very long values
            if len(samples) >= 5:
                break
        cols.append({
            "column_name": col,
            "column_type": str(series.dtype),
            "sample_values": samples,
        })
    return cols


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    # Default to cfg.classify_* where available so the script reflects the
    # currently-configured data source without operator intervention.  See
    # feedback_dynamic_annotations.md: the vocabulary + corpus are
    # runtime-selected; baking values here invites silent drift.
    _cfg = None
    try:
        from atelier.config import load_config
        _cfg = load_config()
    except Exception:
        _cfg = None
    default_connection = (getattr(_cfg, "classify_connection_name", None) or "hive-poc")
    # cfg.classify_database is where the vocab annotations table lives
    # (conventionally "default").  The DATA database (corpus tables) is
    # a separate concept with no canonical cfg key — keep the hardcoded
    # default for now; operators override via --data-database.
    default_database = (getattr(_cfg, "classify_database", None) or "default")
    default_data_database = "reference_corpus"

    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="OPTIONAL operator-review spreadsheet (progressive enhancement). "
             "When provided, adds per-column Atelier-vs-llm-solution context to "
             "the working set (Cat A entries).  When omitted (default), the "
             "script does best-effort curation against the configured Hive "
             "data source.",
    )
    parser.add_argument(
        "--run-id",
        default=None,
        help="Run id under build/results/ to pull current Atelier predictions from. "
             "Defaults to most-recently-modified run dir.  If no runs exist, "
             "the working set is built without Atelier-current context (Cat C only).",
    )
    parser.add_argument(
        "--connection-name",
        default=default_connection,
        help=f"CAI connection name for the vocabulary table (default: {default_connection}).",
    )
    parser.add_argument(
        "--database",
        default=default_database,
        help=f"Hive database holding the 'annotations' vocabulary table (default: {default_database}).",
    )
    parser.add_argument(
        "--data-database",
        default=default_data_database,
        help=f"Hive database holding the corpus tables for curation (default: {default_data_database}).",
    )
    parser.add_argument(
        "--results-dir",
        type=Path,
        default=Path("build/results"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("build/data/agent_mediated/working_set.json"),
    )
    args = parser.parse_args()

    # ── Resolve run id (optional — best-effort if no runs exist) ──
    if args.run_id:
        run_dir = args.results_dir / args.run_id
    else:
        run_dir = _latest_run_dir(args.results_dir)
    if run_dir is not None:
        print(f"Using run dir: {run_dir.name}", file=sys.stderr)
    else:
        print("No prior runs under build/results/ — proceeding without Atelier-current context.", file=sys.stderr)

    # ── Pull inputs ───────────────────────────────────────────────
    print("Loading vocabulary from Hive...", file=sys.stderr)
    vocab = _load_vocabulary(args.connection_name, args.database)
    print(f"  {len(vocab)} vocabulary entries", file=sys.stderr)

    # xlsx is optional (progressive enhancement) — skip cleanly when absent
    if args.xlsx is not None:
        print(f"Parsing xlsx {args.xlsx.name}...", file=sys.stderr)
        xlsx_rows = _parse_xlsx(args.xlsx)
        # Normalize xlsx sheet names to Hive table names (some sheets had
        # typos / spaces since corrected in Hive).
        xlsx_rows = {
            f"{XLSX_NAME_FIXES.get(r['table_name'], r['table_name'])}.{r['column_name']}": {
                **r,
                "table_name": XLSX_NAME_FIXES.get(r['table_name'], r['table_name']),
            }
            for _, r in xlsx_rows.items()
        }
        xlsx_tables = sorted({r["table_name"] for r in xlsx_rows.values()})
        print(f"  {len(xlsx_rows)} column rows across {len(xlsx_tables)} tables", file=sys.stderr)
    else:
        xlsx_rows = {}
        xlsx_tables = []
        print("No xlsx supplied — best-effort curation against the configured Hive data source.", file=sys.stderr)

    if run_dir is not None:
        print(f"Loading classifications.json from run {run_dir.name}...", file=sys.stderr)
        current = _load_classifications(run_dir)
        print(f"  {len(current)} columns in current run", file=sys.stderr)
    else:
        current = {}

    # ── Resolve the full table set under review ──────────────────
    # Hive discovery is always on now (was opt-in via --include-all-hive).
    # The configured data_database is the source of truth; xlsx + run_dir
    # only add per-column context where they overlap.
    print(
        f"Listing tables in Hive {args.data_database}...",
        file=sys.stderr,
    )
    all_hive_tables = set(_list_hive_tables(
        args.connection_name, args.data_database,
    ))
    print(f"  {len(all_hive_tables)} tables in {args.data_database}", file=sys.stderr)

    run_tables = sorted({r["table_name"] for r in current.values()})
    cat_A = set(xlsx_tables) & all_hive_tables  # xlsx + Hive
    cat_B = (set(run_tables) & all_hive_tables) - set(xlsx_tables)  # run only
    cat_C = all_hive_tables - set(xlsx_tables) - set(run_tables)  # neither
    print(
        f"  Cat A (xlsx+Hive): {len(cat_A)}  "
        f"Cat B (run only): {len(cat_B)}  "
        f"Cat C (Hive only): {len(cat_C)}",
        file=sys.stderr,
    )

    # ── Assemble per-column entries ──────────────────────────────
    columns_out = {}

    def _build_atelier_current(cur):
        if cur is None:
            return None
        return {
            "run_id": (run_dir.name if run_dir is not None else None),
            "predicted_annotation": cur.get("predicted_annotation"),
            "predicted_code": str(cur.get("predicted_code") or ""),
            "predicted_label": cur.get("predicted_label"),
            "belief": cur.get("belief"),
            "plausibility": cur.get("plausibility"),
            "conflict": cur.get("conflict"),
            "confidence": cur.get("confidence"),
            "needs_clarification": cur.get("needs_clarification"),
            "evidence_text": cur.get("embedding_text"),
            "evidence_sources": cur.get("evidence_sources"),
            "cautious_code": cur.get("cautious_code"),
        }

    # Cat A: xlsx-driven (full triangulation)
    for key, x in sorted(xlsx_rows.items()):
        cur = current.get(key)
        columns_out[key] = {
            "table_name": x["table_name"],
            "column_name": x["column_name"],
            "category": "A",
            "xlsx": {
                "atelier": x["atelier"],
                "llm_solution": x["llm_solution"],
            },
            "atelier_current": _build_atelier_current(cur),
        }

    # Cat B: run-only (Atelier-current + evidence_text)
    for key, cur in sorted(current.items()):
        if key in columns_out:
            continue
        table = cur.get("table_name")
        if table not in cat_B:
            continue
        columns_out[key] = {
            "table_name": table,
            "column_name": cur.get("column_name"),
            "category": "B",
            "xlsx": None,
            "atelier_current": _build_atelier_current(cur),
        }

    # Cat C: Hive-sample-only (no Atelier signal)
    print(
        f"Sampling {len(cat_C)} Cat C tables from {args.data_database}...",
        file=sys.stderr,
    )
    for t in sorted(cat_C):
        try:
            cols = _sample_hive_columns(
                args.connection_name, args.data_database, t,
            )
            print(f"  {t}: {len(cols)} columns", file=sys.stderr)
        except Exception as exc:
            print(f"  {t}: sample failed — {exc}", file=sys.stderr)
            continue
        for col in cols:
            key = f"{t}.{col['column_name']}"
            columns_out[key] = {
                "table_name": t,
                "column_name": col["column_name"],
                "category": "C",
                "xlsx": None,
                "atelier_current": None,
                "hive_sample": {
                    "column_type": col["column_type"],
                    "sample_values": col["sample_values"],
                },
            }

    tables_present = sorted({c["table_name"] for c in columns_out.values()})

    # ── Write output ─────────────────────────────────────────────
    args.output.parent.mkdir(parents=True, exist_ok=True)
    overlap = set(xlsx_rows) & set(current)
    xlsx_only = set(xlsx_rows) - set(current)
    output = {
        "metadata": {
            "built_at": _utc_iso(),
            "xlsx_path": (str(args.xlsx) if args.xlsx is not None else None),
            "run_id": (run_dir.name if run_dir is not None else None),
            "data_database": args.data_database,
            "vocabulary_size": len(vocab),
            "table_count": len(tables_present),
            "column_count": len(columns_out),
            "categories": {
                "A_xlsx_and_hive": len(cat_A),
                "B_run_only": len(cat_B),
                "C_hive_only": len(cat_C),
            },
            "overlap_with_current_run": len(overlap),
            "xlsx_only_columns": len(xlsx_only),
        },
        "vocabulary": vocab,
        "tables": tables_present,
        "columns": columns_out,
    }
    args.output.write_text(json.dumps(output, indent=2))
    print(f"\nWorking set written: {args.output}", file=sys.stderr)
    print(
        f"  {len(columns_out)} columns across {len(tables_present)} tables, "
        f"{len(vocab)} vocab entries "
        f"(Cat A={len(cat_A)}, Cat B={len(cat_B)}, Cat C={len(cat_C)})",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
